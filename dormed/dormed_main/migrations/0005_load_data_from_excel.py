
from django.db import migrations, models
from openpyxl import load_workbook
from dormed_main.models import Price

def load_data(apps, dormed_main_price):
    wb = load_workbook(filename='C:/IT/dormed-django/table.xlsx')
    ws = wb.active
    for row in ws.iter_rows(min_row = 1):
        price = Price(
            item = row[0].value,
            price_dormed = row[1].value,
            price_others= row[2].value,
            )
        price.save()

class Migration(migrations.Migration):
    
    dependencies = [
        ('dormed_main', '0004_remove_price_description_remove_price_price_and_more')
        ]
    operations = [
        migrations.RunPython(load_data),
    ]